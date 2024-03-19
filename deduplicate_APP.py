import sys
from functools import partial
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from PySide6.QtCore import QRunnable, Qt, QThreadPool, Signal
from PySide6.QtGui import QAction, QDragEnterEvent, QDropEvent, QScreen
from PySide6.QtWidgets import (QApplication, QCheckBox, QHBoxLayout, QLabel,
                               QMenu, QMessageBox, QProgressBar, QPushButton,
                               QTreeWidget, QTreeWidgetItem, QVBoxLayout,
                               QWidget)


class MyTask(QRunnable):
    def __init__(self, function, *args):
        super().__init__()

    def run(self):
        """
        重写 run 方法
        :return: 无
        """


class MainWidget(QWidget):
    def __init__(self):
        """
        初始化
        :return: 无
        """
        super().__init__()

        self.file_infos = {}  # 保存文件信息
        self.checkbox_infos = {}  # 保存 QCheckBox 的信息
        self.progressBar_value = 0  # 进度条的值

        self.max_cols = 0  # 最大列数

        # 线程池
        self.thread_pool = QThreadPool()
        self.thread_pool.setMaxThreadCount(4)

        self.initUI()

    def initUI(self):
        """
        初始化UI
        :return: 无
        """
        self.setAcceptDrops(True)

        layout = QVBoxLayout()
        checkbox_layout = QHBoxLayout()

        self.label = QLabel("将 Excel 文件或文件夹拖放到框内:", self)
        self.treeWidget = QTreeWidget()
        self.btn_clear = QPushButton('清空列表')

        self.label_tips = QLabel('请选择要检测重复的列号:')

        self.widget = QWidget()

        for i in range(1, 14):
            # 新建 QCheckBox
            checkBox = QCheckBox()
            checkBox.setText(f'第{i}列')

            # checkBox 状态改变信号处理
            checkBox.stateChanged.connect(partial(self.on_checkBox_state_changed, index=i))

            # 将 checkBox 添加到水平布局
            checkbox_layout.addWidget(checkBox)

            # 将编号与 QCheckBox 对应起来
            self.checkbox_infos[i] = {
                'obj': checkBox,
                'checkState': False
            }

        self.widget.setLayout(checkbox_layout)

        self.btn_cancel = QPushButton('停止执行')
        self.btn = QPushButton('开始去重')

        self.label_result = QLabel('...')
        self.progressBar = QProgressBar()

        self.contex_menu = QMenu(self)  # 创建菜单

        layout.addWidget(self.label)
        layout.addWidget(self.treeWidget)
        layout.addWidget(self.btn_clear)

        layout.addWidget(self.label_tips)
        layout.addWidget(self.widget)

        layout.addWidget(self.btn_cancel)
        layout.addWidget(self.btn)
        layout.addWidget(self.label_result)

        layout.addWidget(self.progressBar)

        self.setLayout(layout)

        # 为 TreeWidget 添加列表头
        self.treeWidget.setHeaderLabels(['文件名', '是否有重复项', '总行数', '总列数', '重复数据行数'])
        self.treeWidget.header().setDefaultAlignment(Qt.AlignCenter)  # 居中显示

        # 设置 TreeWidget 可拖拽
        self.treeWidget.setAcceptDrops(True)
        self.treeWidget.setDragEnabled(True)

        # 为右键菜单添加移除选项
        self.remove_action = QAction("移除选项", self, triggered=self.removeItem)
        self.contex_menu.addAction(self.remove_action)

        # 为 QTreeWidget 绑定右键菜单
        self.treeWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.treeWidget.customContextMenuRequested.connect(self.showContexMenu)

        self.btn_clear.clicked.connect(self.clearList)  # 清除列表的按键点击信号绑定
        self.btn_cancel.clicked.connect(self.cancelProcess_excel)
        self.btn.clicked.connect(self.process_excel)  # 开始去重按钮点击信号绑定

        self.setWindowTitle('Excel 去重软件')  # 窗体标题
        self.setWindowOpacity(0.95)

        self.setGeometry(800, 500, 890, 550)  # x, y, w, h

        self.centerOnScreen()  # 居中显示窗体

    def centerOnScreen(self):
        """
        将窗体移动到屏幕中央
        :return: 无
        """
        rect = self.frameGeometry()
        centerPoint = QScreen.availableGeometry(QApplication.primaryScreen()).center()
        rect.moveCenter(centerPoint)
        self.move(rect.topLeft())

    def dragEnterEvent(self, event: QDragEnterEvent):
        """
        拖拽进入事件
        :param event: 事件
        :return: 无
        """
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        """
        拖拽事件
        :param event: 事件
        :return: 无
        """
        for url in event.mimeData().urls():
            filepath = url.toLocalFile()  # 获取文件路径
            fp = Path(filepath)
            # 用户拖入的是文件夹
            if fp.is_dir():
                # 文件后缀是 .xlsx 并且尚未拖入的
                for f in fp.glob('**/*.xlsx'):
                    if not self.file_infos.get(filepath):
                        self.addFilePath_to_TreeView(str(f))    # 将文件路径显示到 TreeView 中

                # 文件后缀是 .xls 并且尚未拖入的
                for f in fp.glob('**/*.xls'):
                    if not self.file_infos.get(filepath):
                        self.addFilePath_to_TreeView(str(f))     # 将文件路径显示到 TreeView 中

            # 用户拖入的是文件
            elif not self.file_infos.get(filepath) and (filepath.endswith('.xlsx') or filepath.endswith('.xls')):
                self.addFilePath_to_TreeView(filepath)           # 将文件路径显示到 TreeView 中

    def addFilePath_to_TreeView(self, filepath: str):
        """
        添加文件路径到 TreeView 中
        :param filepath: 文件路径
        :return: 无
        """
        item = QTreeWidgetItem(self.treeWidget)     # 新建根节点
        item.setText(0, filepath)                   # 设置根节点的文本

        item.setTextAlignment(1, Qt.AlignCenter)    # 设置根节点第2列的文本居中
        item.setTextAlignment(2, Qt.AlignCenter)    # 设置根节点第3列的文本居中
        item.setTextAlignment(3, Qt.AlignCenter)    # 设置根节点第4列的文本居中
        item.setTextAlignment(4, Qt.AlignCenter)    # 设置根节点第5列的文本居中

        self.treeWidget.resizeColumnToContents(0)   # 根据列中的内容, 自动设置第1列的宽度
        self.treeWidget.resizeColumnToContents(1)   # 根据列中的内容, 自动设置第2列的宽度
        self.treeWidget.resizeColumnToContents(2)   # 根据列中的内容, 自动设置第3列的宽度
        self.treeWidget.resizeColumnToContents(3)   # 根据列中的内容, 自动设置第4列的宽度
        self.treeWidget.resizeColumnToContents(4)   # 根据列中的内容, 自动设置第5列的宽度

        # 获取该节点的 index
        index = self.treeWidget.indexOfTopLevelItem(item)

        # 获取该文件的总行数、总列数
        rowCol = self.getExcelFileRowsCols(filepath)
        max_rows = str(rowCol[0])
        self.max_cols = max_cols = rowCol[1]

        # 该文本对应的 QTreeWidgetItem 的index、是否有重复数据、总行数、总列数、重复行数
        self.file_infos[filepath] = {'index': index, 'isRepeat': False, 'max_rows': max_rows, 'max_cols': max_cols,
                                     'repeat_rows': '0'}

        # 在 TreeView 中显示文件的总行数、总列数
        self.updateTreeView(filepath, True)

    def updateTreeView(self, filepath: str, only_rowcol: bool):
        """
        更新 QTreeWidget
        :param filepath: 文件路径
        :param max_rows: 文件总行数
        :return: 无
        """
        index = self.file_infos[filepath]['index']      # 在字典中查找控件的 index
        item = self.treeWidget.topLevelItem(int(index)) # 根据 index 获取 QTreeWidgetItem 对象

        max_rows = self.file_infos.get(filepath).get('max_rows') # 获取文件的总行数
        max_cols = self.file_infos.get(filepath).get('max_cols') # 获取文件的总列数

        # 只更新文件总行数和总列数
        if only_rowcol:
            item.setText(2, str(max_rows))
            item.setText(3, str(max_cols))
            return

        isRepeat = self.file_infos.get(filepath).get('isRepeat')
        repeat_rows = self.file_infos.get(filepath).get('repeat_rows')

        if isRepeat:
            item.setText(1, '是')
        else:
            item.setText(1, '否')

        item.setText(4, str(repeat_rows))

    def showContexMenu(self, point):
        """
        将右键菜单显示在屏幕中
        :param point: 坐标
        :return: 无
        """
        # 要显示菜单的坐标
        showPoint = self.mapToGlobal(point)
        showPoint.setX(showPoint.x() + 35)
        showPoint.setY(showPoint.y() + 55)

        # 在指定位置处显示
        self.contex_menu.exec(showPoint)

    def getExcelFileRowsCols(self, filepath) -> Tuple:
        """
        获取 Excel 文件的总行数
        :param filepath: 文件路径
        :return: 无
        """
        wb = load_workbook(filepath)
        ws = wb.active

        return ws.max_row, ws.max_column

    def removeItem(self):
        """
        移除 QTreeWidget 中的项目
        :return: 无
        """
        for selected_item in self.treeWidget.selectedItems():
            index = self.treeWidget.indexOfTopLevelItem(selected_item)  # 获取选中节点的 index

            self.treeWidget.takeTopLevelItem(index)     # 因为是根节点, 所以直接从根节点中移除
            self.file_infos.pop(selected_item.text(0))  # 从 file_infos 中移除

            # 前面删掉一个, 后面的 index 就得 -1
            for i in self.file_infos:
                if self.file_infos[i]['index'] > index:
                    self.file_infos[i]['index'] -= 1

    def updateProgressBar(self):
        """
        更新进度条
        :return: 无
        """

    def setLable(self, label: QLabel, text, color='#000'):
        """
        设定 QLabel 的文本和样式
        :param label: QLabel 对象
        :param text: label 上要显示的文本
        :param color: label 文字上的颜色
        :return: 无
        """
        if label is None:
            return
        label.setText(text)
        label.setStyleSheet(f'color: {color}')

    def on_checkBox_state_changed(self, checked, index):
        """
        QCheckBox 状态改变时要执行的槽函数
        :param checked: > 0 为选中; 其余为未选中
        :param index: 选中的 QCheckBox 的编号
        :return: 无
        """
        self.checkbox_infos[index]['checkState'] = int(checked) > 0

    def getAllCheckBoxState(self) -> Dict:
        """
        获取所有 checkBox 的状态, 判断用户是否勾选了 checkBox
        :return:  返回一个字典. {1: False, 1: True, ...}
        """
        res_dict = {
            key: self.checkbox_infos[int(key)]['checkState']
            for key in self.checkbox_infos
            if int(key) <= self.max_cols
        }
        return res_dict

    def clearList(self):
        """
        清空文件列表
        :return: 无
        """
        self.treeWidget.clear()
        self.file_infos.clear()
        self.setLable(self.label_result, '...')

    def cancelProcess_excel(self):
        """
        TODO: 取消处理 Excel 文件
        :return: 无
        """

    def process_excel(self):
        """
        处理 Excel 文件
        :return: 无
        """
        # 获取 ListWidget 里面的所有项
        filepath_list = [self.treeWidget.topLevelItem(i).text(0) for i in range(self.treeWidget.topLevelItemCount())]

        # ListWidget 内容为空, 弹窗提示用户并结束函数
        if not filepath_list:
            msgBox = QMessageBox()
            msgBox.setText("请把 Excel 文件拖进来")
            msgBox.setWindowTitle("警告")
            msgBox.setIcon(QMessageBox.Warning)
            msgBox.setStandardButtons(QMessageBox.Ok)
            msgBox.exec()
            return

        res_dict = self.getAllCheckBoxState()       # 获取所有 checkBox 的状态
        res_list = [res_dict[i] for i in res_dict]  # 获取所有 checkBox 的状态的布尔值

        # 如果所有 checkBox 都没有勾选
        if not all(res_list):
            ...

        self.btn_clear.setEnabled(False)    # 禁用清空列表按钮
        self.btn.setEnabled(False)          # 禁用开始去重按钮

        self.setLable(self.label_result, '处理中, 请您稍等...')

        res_list = []
        for filepath in filepath_list:
            # 获取要用哪几列来检查重复
            cols = [int(i) for i in res_dict if res_dict[i]]

            # 调用去重函数, 并将返回结果存放在 res_list 列表中
            res = self.deduplicate_excel(Path(filepath), cols)
            res_list.append(res)

            self.updateTreeView(str(filepath), False)  # 更新 TreeWidget 信息

        if any(res_list):
            self.setLable(self.label_result, '处理完成。 结果在 "去重结果" 文件夹中, 若没有这个文件夹说明没有重复项',
                          '#00b440')
        else:
            self.setLable(self.label_result, '该列表的 Excel 文件都没有重复项~', '#ffb434')

        self.btn_clear.setEnabled(True) # 启用清空列表按钮
        self.btn.setEnabled(True)       # 启用开始去重按钮

    def deduplicate_excel(self, filepath: Path, cols: List[int]) -> bool:
        """
        Excel 文件去重的核心函数
        :param filepath: Excel 文件的路径
        :param cols: 要用哪几列来检查重复
        :return: 返回 bool 值. 如果有重复的就返回 True, 没有重复数据就返回 False
        """


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWidget()
    window.show()
    sys.exit(app.exec())
